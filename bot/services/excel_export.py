"""
Сервис формирования Excel-отчёта по товарам пользователя.
Список полей и их подписи заданы в bot/keyboards/inline.py (EXCEL_FIELDS),
чтобы кнопки выбора полей и колонки в файле всегда совпадали.
"""

import os
import tempfile
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from bot.database.models import Product
from bot.keyboards.inline import EXCEL_FIELDS


def _field_value(product: Product, key: str):
    if key == "name":
        return product.name or f"Артикул {product.nm_id}"
    if key == "article":
        return product.nm_id
    if key == "price":
        return product.current_price
    if key == "old_price":
        return product.price_without_discount
    if key == "brand":
        return product.brand or "—"
    if key == "rating":
        return product.rating
    if key == "feedbacks":
        return product.feedbacks_count
    if key == "seller":
        return product.seller or "—"
    if key == "url":
        return product.url or f"https://www.wildberries.ru/catalog/{product.nm_id}/detail.aspx"
    return ""


def build_excel_report(products: List[Product], field_keys: List[str]) -> str:
    """
    Строит .xlsx файл с выбранными товарами и выбранными полями.
    Возвращает путь к временному файлу.
    """
    # Порядок колонок всегда как в EXCEL_FIELDS, а не в том порядке, в котором пользователь тыкал кнопки
    ordered_fields = [(k, label) for k, label in EXCEL_FIELDS if k in field_keys]

    wb = Workbook()
    ws = wb.active
    ws.title = "WB Мониторинг"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    for col_idx, (key, label) in enumerate(ordered_fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label.split(" ", 1)[-1])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, product in enumerate(products, start=2):
        for col_idx, (key, _label) in enumerate(ordered_fields, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_field_value(product, key))

    # Автоширина колонок (по самому длинному значению в столбце, с разумным потолком)
    for col_idx, (key, label) in enumerate(ordered_fields, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(label.split(" ", 1)[-1])
        for row_idx in range(2, len(products) + 2):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    ws.freeze_panes = "A2"

    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="wb_report_")
    os.close(fd)
    wb.save(path)
    return path
